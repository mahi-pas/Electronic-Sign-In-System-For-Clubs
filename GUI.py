from tkinter import * 
import os
import openpyxl
from openpyxl import Workbook
from datetime import date
from tkinter import messagebox
import ctypes

class GUI:
    def __init__(self, master, size = "925x510"):
        self.master = master
        self.master.title("Sign In")
        self.master.geometry(size) #GUI start size

        #set up workbook
        self.wb = openpyxl.load_workbook('data.xlsx')
        self.ws = self.wb.active

        #check for if today's column exists, if no make it
        self.currentcol=self.ws.cell(row = 1, column=self.ws.max_column).value
        self.today = str(date.today())
        if (self.currentcol==self.today):
            self.signin_col=self.ws.max_column
        else:
            self.ws.cell(row=1,column=(self.ws.max_column+1)).value = self.today
            self.signin_col=self.ws.max_column
            self.wb.save('data.xlsx')

        #label
        self.label = Label(master, text="Enter Full Name", font="Helvetica 36 bold")
        self.label.pack()

        #input box
        self.inputbox = Entry(self.master, font = "Helvetica 44 bold", width=25)
        self.inputbox.pack()

        #signin button
        self.signin_button = Button(self.master, font = "Helvetica 36 bold", text="Sign In", command=self.signin)
        self.signin_button.pack()

    #functions

    #messagebox
    def Mbox(self, title, text, style):
        return ctypes.windll.user32.MessageBoxW(0, text, title, style)

    #signin
    def signin(self):
        self.person = self.inputbox.get()
        self.inputbox.delete(0, 'end')
        #setting macros i guess(just makes it easier)
        self.name_col = 2
        self.name_row = self.ws.max_row+1

        self.found = False #variable for if name is found
        #finding name in list and checking them off
        for row in range(2,self.ws.max_row+1):   
            if (self.person.capitalize() == (self.ws.cell(row=row,column=self.name_col).value).capitalize()): #checks if the same
                self.ws.cell(row=row,column=self.signin_col).value = "Present"
                self.wb.save('data.xlsx')
                self.found = True
                self.Mbox('Sign In', 'You have successfully signed in', 0)
                break

        #if name isn't found ask if they are new or mispelled
        if (not self.found): 
            message_answer = self.Mbox('New Member Prompt', 'Name not found. Are you a new member?', 4)
                #6 = yes
                #7 = no
            if message_answer == 6: #yes
                #code to add a new name at the bottom
                self.ws.cell(row=(self.ws.max_row+1),column=self.name_col).value = self.person
                self.ws.cell(row=self.ws.max_row,column=self.signin_col).value = "Present"
                self.wb.save('data.xlsx')
            elif message_answer == 7: #no
                self.Mbox('New Member Prompt', 'Your name was typed incorrectly. Retype it.', 0)






